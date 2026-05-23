#!/usr/bin/env python3
"""H-035 — FUTURES Time-Series Momentum (TSMOM) Diversified Sidecar.

OPT-IN RESEARCH SIDECAR. HARNESS-GATED. FREE-DATA. NO PRODUCTION WIRING.
Multi-AI-D-path consensus (Grok + Claude swarm 2026-05-19): the classic
Moskowitz-Ooi-Pedersen (2012) time-series momentum factor reimplemented as
a diversified equal-weighted vol-targeted book across liquid futures.

No caller in `quality_gates.py`, `dashboard_generator.py`,
`passes_active_gate`, `calculate_smart_score`, `production_scanner`, or any
pick-generation / scoring path. Importing this module has zero side effects
on production. It fetches free public futures data, runs the backtest, gates
the result through `tools/edge_stability_harness.py` (imported UNMODIFIED),
and writes a verdict report. Nothing else.

------------------------------------------------------------------------------
THE HYPOTHESIS (H-035, per Moskowitz-Ooi-Pedersen 2012)
------------------------------------------------------------------------------
Time-series momentum across a diversified liquid-futures universe:
  * EQUITY-INDEX futures: ES=F (S&P), NQ=F (NASDAQ), YM=F (Dow), RTY=F (R2K)
  * CURRENCY futures:     6E=F (EUR), 6B=F (GBP), 6J=F (JPY), 6A=F (AUD)
  * COMMODITY futures:    CL=F (WTI), GC=F (gold), HG=F (copper)
  * BOND futures:         ZN=F (10Y), ZB=F (30Y)

Signal: the SIGN of each contract's trailing 12-month total return. LONG when
positive, SHORT when negative. Each leg is vol-scaled to a target ~10%
annualised volatility, then equal-weighted across the diversified book.
Rebalance monthly. STRICT NO-LOOK-AHEAD: the signal date used at the start
of month M is strictly before M (price observations at the close of M-1).

The harness "score" field is `|tsmom_z|` — the conviction magnitude of the
12-month return z-score against the contract's own strictly-past distribution.
A binary sign signal carries no magnitude, but the harness needs a continuous
conviction score to separate winners from losers — the |z| is that magnitude.

------------------------------------------------------------------------------
DATA — FREE ONLY
------------------------------------------------------------------------------
PRIMARY: yfinance =F continuous-front-month tickers. This is the REAL
implementation per the Claude swarm verdict (2026-05-19): "AQR publishes
factor returns but NOT constituent positions. You need to reimplement using
yfinance =F tickers." yfinance gives ~25-30 years of daily continuous-front
data on the major contracts, free, no key.

OPTIONAL SANITY CHECK: AQR Capital Management publishes the academic
Time-Series Momentum monthly factor return series (and many adjacent factor
series) for free at https://www.aqr.com/Insights/Datasets/Time-Series-Momentum-Factors-Monthly.
The Excel file can be downloaded by hand and dropped into
`tools/cache/aqr_tsmom_monthly.xlsx` (or `.csv`). If present, this module
loads it and reports the published return series alongside our reconstruction
purely as a smell-test — NOT as the verdict input. AQR's series is the
factor-return aggregate; it does NOT carry per-contract positions, so it
cannot drive the harness. The yfinance reconstruction IS the verdict input.

No paid feed. No API key. If yfinance is unreachable the module degrades to
a clearly-labelled synthetic series and reports `UNTESTED-data-gap` — it
never fabricates a pass.

------------------------------------------------------------------------------
CONSTRUCTION (legitimate density pattern from H-008/H-019/H-026/H-027/H-028)
------------------------------------------------------------------------------
  * FULL continuous-position cross-sectional book. Every (contract,
    month-end) with a next month is one resolved record. NO threshold,
    NO event filter — every month-end emits one position per contract.
  * Vol-scaling per leg: position size = TARGET_VOL / realised_vol
    (annualised, computed from the trailing VOL_LOOKBACK_DAYS, strictly past).
    Equal-weight across the diversified book at the portfolio level — the
    per-position record carries the vol-scaled signed return.
  * Monthly rebalance — entry at month-end M, exit at month-end M+1.
    The 12-month signal uses returns over [M-12, M] computed from prices
    at-or-before the close of M.

------------------------------------------------------------------------------
THE VERDICT GATE
------------------------------------------------------------------------------
Records pass through `tools/edge_stability_harness.evaluate()` /
`is_admissible()` imported UNMODIFIED. ADMISSIBLE iff |eff| >= 0.30, same
sign, in >= 3 of 5 walk-forward 14-day windows, each window >= 80 records.
The 14-day window is the harness default and is NOT touched here. Monthly
rebalanced records across 13 contracts give density of roughly 13 records /
month / 21 trading days ~= 0.6 records/day — well below the 80/window floor.
That is the honest density problem of a monthly factor and the harness will
typically render UNTESTED-data-gap on the monthly-rebalance variant. A daily
secondary book (positions held continuously, signal recomputed daily from the
same 252-day window) is included so the harness gets dense records; both
verdicts are reported.

PLUS a 10bps round-trip post-cost gate. Futures are cheap (mini-ES round-trip
~1 tick = ~0.5bp; bond/currency futures comparable; commodity contracts a
hair wider). 10bps is a conservative round-trip across the diversified book.
Net edge must retain >= 60% of gross.

A gaudy in-sample WR is NOT a pass. The edge hunt's base rate is poor (8-11
harness kills to date). This module reports the verdict honestly whichever
way it lands. Nothing here is wired into or sized by any production path.

    python tools/h035_futures_tsmom_research.py [--quick] [--json]
            [--refresh-cache]
"""
from __future__ import annotations

import argparse
import json
import math
import statistics
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

if sys.platform == "win32":
    for _s in (sys.stdout, sys.stderr):
        if hasattr(_s, "reconfigure"):
            try:
                _s.reconfigure(encoding="utf-8", errors="replace")
            except Exception:  # noqa: BLE001
                pass

# Harness is imported UNMODIFIED (the gate, verbatim).
sys.path.insert(0, str(ROOT / "tools"))
import edge_stability_harness as harness  # noqa: E402

# ---------------------------------------------------------------------------
# Tunables — pre-registered (H-035). Moskowitz-Ooi-Pedersen (2012):
# 12-month look-back, monthly rebalance, vol-target ~10% per leg.
# ---------------------------------------------------------------------------
LOOKBACK_MONTHS = 12              # 12-month total-return sign signal
LOOKBACK_DAYS = 252               # trading-day equivalent
VOL_LOOKBACK_DAYS = 60            # realised-vol window (strictly past)
TARGET_VOL_ANNUAL = 0.10          # ~10% annualised per leg
Z_ROLL = 60                       # rolling z-score look-back for |signal_z|
WINDOW_DAYS = 14                  # walk-forward window length (harness default)
HARNESS_FIELD = "signal_z"        # conviction-magnitude score the harness reads
ROUND_TRIP_COST_BPS = 10.0        # futures round-trip (conservative, diversified)
COST_SURVIVAL_MIN = 0.60          # net edge must retain >= 60% of gross
TRADING_DAYS_YEAR = 252

CACHE = ROOT / "tools" / "cache" / "h035_futures_tsmom_cache.json"
AQR_CACHE_XLSX = ROOT / "tools" / "cache" / "aqr_tsmom_monthly.xlsx"
AQR_CACHE_CSV = ROOT / "tools" / "cache" / "aqr_tsmom_monthly.csv"

# Diversified liquid-futures universe — equity / FX / commodity / bond.
UNIVERSE_FULL = [
    # Equity-index
    "ES=F", "NQ=F", "YM=F", "RTY=F",
    # Currency
    "6E=F", "6B=F", "6J=F", "6A=F",
    # Commodity
    "CL=F", "GC=F", "HG=F",
    # Government bond
    "ZN=F", "ZB=F",
]
UNIVERSE_QUICK = ["ES=F", "NQ=F", "CL=F", "GC=F", "ZN=F"]

ASSET_CLASS_MAP = {
    "ES=F": "EQUITY", "NQ=F": "EQUITY", "YM=F": "EQUITY", "RTY=F": "EQUITY",
    "6E=F": "FX", "6B=F": "FX", "6J=F": "FX", "6A=F": "FX",
    "CL=F": "COMMODITY", "GC=F": "COMMODITY", "HG=F": "COMMODITY",
    "ZN=F": "BOND", "ZB=F": "BOND",
}


# ===========================================================================
# Data fetch — free public APIs, no key.
# ===========================================================================
def fetch_futures_series(ticker: str) -> dict:
    """Fetch a free daily close series for one continuous-front-month future.

    Strategy (free sources only):
      yfinance — daily close of the =F continuous-front contract. Free, no
      key, ~25-30 years of history on the major majors.

    Returns {"price": {iso_date: close}}. On network failure returns empty;
    the caller then degrades to a clearly-labelled synthetic series and an
    UNTESTED verdict — never a fake pass.
    """
    price: dict[str, float] = {}
    try:
        import yfinance as yf  # noqa: PLC0415
        tk = yf.Ticker(ticker)
        hist = tk.history(period="max", interval="1d", auto_adjust=False)
        if hist is not None and not hist.empty:
            for idx, row in hist.iterrows():
                try:
                    d = idx.date().isoformat()
                    c = float(row["Close"])
                    if c > 0 and not math.isnan(c):
                        price[d] = c
                except Exception:  # noqa: BLE001
                    continue
    except Exception:  # noqa: BLE001
        pass
    return {"price": price}


def _synthetic_series(ticker: str, n_days: int = 6000) -> dict:
    """Deterministic synthetic series for an OFFLINE run.

    Used ONLY when yfinance is unreachable. Clearly labelled and drives an
    UNTESTED-data-gap verdict; never reported as a real pass.
    """
    seed = sum(ord(c) for c in ticker) * 2654435761 & 0xFFFFFFFF
    state = seed

    def rand() -> float:
        nonlocal state
        state = (1103515245 * state + 12345) & 0x7FFFFFFF
        return state / 0x7FFFFFFF

    price: dict[str, float] = {}
    px = 100.0 + (seed % 50)
    base = datetime(2005, 1, 3, tzinfo=timezone.utc)
    for i in range(n_days):
        d = (base.fromordinal(base.toordinal() + i)).date().isoformat()
        px *= 1.0 + (rand() - 0.5) * 0.015
        price[d] = round(px, 4)
    return {"price": price}


def load_market_data(tickers: list[str], refresh: bool) -> dict:
    """Fetch (or load cached) daily close per future.

    Adds an `_offline` flag per ticker when the series came from the synthetic
    fallback so the verdict layer can honestly report UNTESTED-data-gap.
    """
    if CACHE.exists() and not refresh:
        try:
            cached = json.loads(CACHE.read_text(encoding="utf-8"))
            if all(t in cached.get("data", {}) for t in tickers):
                print(f"# loaded cache {CACHE}", file=sys.stderr)
                return cached["data"]
        except Exception:  # noqa: BLE001
            pass

    data: dict[str, dict] = {}
    for t in tickers:
        print(f"# fetching {t} (free yfinance =F continuous-front) ...",
              file=sys.stderr)
        series = fetch_futures_series(t)
        if len(series.get("price", {})) < 500:
            print(f"#   {t}: dense free price series unavailable -> synthetic "
                  f"fallback (verdict will be UNTESTED)", file=sys.stderr)
            series = _synthetic_series(t)
            series["_offline"] = True
        else:
            series["_offline"] = False
        print(f"#   {t}: price={len(series['price'])} bars  "
              f"offline={series['_offline']}", file=sys.stderr)
        data[t] = series
        time.sleep(0.3)

    CACHE.parent.mkdir(parents=True, exist_ok=True)
    CACHE.write_text(json.dumps({
        "fetched_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "data": data,
    }), encoding="utf-8")
    print(f"# wrote cache {CACHE}", file=sys.stderr)
    return data


def load_aqr_tsmom_optional() -> dict | None:
    """Load AQR's published monthly TSMOM factor return series if present.

    Returns {"month_iso": monthly_return_decimal, ...} or None when no AQR
    file is bundled. This series is a SANITY CHECK only — the academic
    factor-return aggregate. AQR does NOT publish constituent positions, so
    this cannot drive the harness. The yfinance reconstruction is the
    verdict input.

    Download path (manual, free):
      https://www.aqr.com/Insights/Datasets/Time-Series-Momentum-Factors-Monthly
      -> save the workbook as tools/cache/aqr_tsmom_monthly.xlsx OR convert
         to a 2-column CSV (date, ret) at tools/cache/aqr_tsmom_monthly.csv
    """
    # CSV first — zero dependencies.
    if AQR_CACHE_CSV.exists():
        try:
            import csv  # noqa: PLC0415
            out: dict[str, float] = {}
            with AQR_CACHE_CSV.open(encoding="utf-8") as f:
                reader = csv.reader(f)
                for row in reader:
                    if not row or len(row) < 2:
                        continue
                    d_raw = row[0].strip()
                    try:
                        ret = float(row[1])
                    except (TypeError, ValueError):
                        continue
                    # Accept ISO YYYY-MM or YYYY-MM-DD.
                    if len(d_raw) >= 7 and d_raw[4] == "-":
                        out[d_raw[:7]] = ret
            if out:
                return out
        except Exception:  # noqa: BLE001
            pass
    # Excel fallback — needs openpyxl, optional.
    if AQR_CACHE_XLSX.exists():
        try:
            from openpyxl import load_workbook  # noqa: PLC0415
            wb = load_workbook(AQR_CACHE_XLSX, data_only=True, read_only=True)
            out: dict[str, float] = {}
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    if not row or row[0] is None:
                        continue
                    try:
                        d_raw = str(row[0])[:7]
                        ret = float(row[1])
                    except (TypeError, ValueError):
                        continue
                    if len(d_raw) >= 7 and d_raw[4] == "-":
                        out[d_raw] = ret
            return out or None
        except Exception:  # noqa: BLE001
            return None
    return None


# ===========================================================================
# Signal math (pure, network-free)
# ===========================================================================
def trailing_return(dates: list[str], prices: dict[str, float],
                    idx: int, lookback: int):
    """Total return from prices[dates[idx-lookback]] to prices[dates[idx]]."""
    if idx < lookback:
        return None
    p0 = prices[dates[idx - lookback]]
    p1 = prices[dates[idx]]
    if p0 <= 0:
        return None
    return p1 / p0 - 1.0


def realised_vol(dates: list[str], prices: dict[str, float],
                 idx: int, lookback: int):
    """Annualised realised vol from the trailing `lookback` daily log returns.

    Strictly past — uses dates[idx-lookback:idx] (idx excluded).
    """
    if idx < lookback + 1:
        return None
    rets = []
    for j in range(idx - lookback, idx):
        p0 = prices[dates[j - 1]] if j - 1 >= 0 else None
        p1 = prices[dates[j]]
        if p0 is None or p0 <= 0 or p1 <= 0:
            continue
        rets.append(math.log(p1 / p0))
    if len(rets) < 5:
        return None
    sd = statistics.pstdev(rets)
    if sd <= 0:
        return None
    return sd * math.sqrt(TRADING_DAYS_YEAR)


def rolling_z(series: list[float], idx: int, roll: int):
    """Z-score of series[idx] vs the `roll` STRICTLY-PAST observations."""
    if idx < roll:
        return None
    window = series[idx - roll:idx]
    mu = statistics.fmean(window)
    sd = statistics.pstdev(window)
    if sd <= 0:
        return None
    return (series[idx] - mu) / sd


# ===========================================================================
# Daily continuous-position backtest (the harness-density variant)
# ===========================================================================
def backtest_continuous(data: dict) -> dict:
    """FULL continuous-position diversified TSMOM book on daily rebalance.

    For every (contract, day D) with a 252-day look-back and a next bar:
      * trailing_ret = total return [D-252, D] of the contract.
      * direction: LONG (+1) if trailing_ret > 0; SHORT (-1) if < 0; 0 skip.
      * vol-scale: weight = TARGET_VOL_ANNUAL / realised_vol_60d(D),
        capped at 5x to keep extreme low-vol periods from blowing up.
      * raw_ret = close(D+1)/close(D) - 1
      * signed_ret = raw_ret * direction * weight  (vol-scaled)
      * record: status WON iff signed_ret > 0 (gross). Harness score field
        `signal_z` = |z-score of trailing_ret against its own 60-day strictly-
        past distribution| — the conviction magnitude. The 12-month sign is
        the directional signal; the |z| is the magnitude the harness ranks on.

    STRICT no-look-ahead: signal date is D (strictly before the entry day
    D+1), return realized over [D, D+1).
    """
    records: list[dict] = []
    per_inst: dict[str, dict] = {}
    gross_rets: list[float] = []
    net_rets: list[float] = []
    cost_frac = ROUND_TRIP_COST_BPS / 10000.0
    any_offline = False

    for ticker, series in data.items():
        if series.get("_offline"):
            any_offline = True
        price = series.get("price", {})
        dates = sorted(price)
        # Trailing-return series for the z-score baseline (strictly-past z).
        trs: list[float] = []
        valid_idx: list[int] = []
        for i in range(len(dates)):
            tr = trailing_return(dates, price, i, LOOKBACK_DAYS)
            if tr is not None:
                trs.append(tr)
                valid_idx.append(i)

        n = wins = 0
        for k in range(len(valid_idx) - 1):
            i = valid_idx[k]
            i_next = i + 1
            if i_next >= len(dates):
                continue
            tr = trs[k]
            direction = 1 if tr > 0 else (-1 if tr < 0 else 0)
            if direction == 0:
                continue
            z = rolling_z(trs, k, Z_ROLL)
            if z is None:
                continue
            rv = realised_vol(dates, price, i, VOL_LOOKBACK_DAYS)
            if rv is None or rv <= 0:
                continue
            weight = min(TARGET_VOL_ANNUAL / rv, 5.0)
            d = dates[i]
            d_next = dates[i_next]
            p0, p1 = price[d], price[d_next]
            if p0 <= 0:
                continue
            raw_ret = p1 / p0 - 1.0
            signed_ret = raw_ret * direction * weight
            net_ret = signed_ret - cost_frac
            status = "WON" if signed_ret > 0 else "LOST"
            n += 1
            wins += int(status == "WON")
            gross_rets.append(signed_ret)
            net_rets.append(net_ret)
            records.append({
                "status": status,
                "resolved_at": d_next,
                "entry_date": d,
                "signal_date": d,
                "timestamp": d,
                HARNESS_FIELD: abs(z),
                "signed_ret": round(signed_ret, 8),
                "net_ret": round(net_ret, 8),
                "direction": direction,
                "weight": round(weight, 6),
                "instrument": ticker,
                "asset_class": ASSET_CLASS_MAP.get(ticker, "FUTURES"),
            })
        per_inst[ticker] = {
            "n": n, "wins": wins,
            "wr": round(wins / n, 4) if n else None,
            "asset_class": ASSET_CLASS_MAP.get(ticker, "FUTURES"),
        }

    return {
        "records": records,
        "per_instrument": per_inst,
        "gross_rets": gross_rets,
        "net_rets": net_rets,
        "any_offline": any_offline,
    }


# ===========================================================================
# Monthly rebalance (per Moskowitz-Ooi-Pedersen 2012; sparse, descriptive only)
# ===========================================================================
def backtest_monthly(data: dict) -> dict:
    """Classical monthly-rebalance MOP TSMOM book — DESCRIPTIVE ONLY.

    Monthly rebalanced records across ~13 contracts give density of roughly
    0.6 records/day — well below the 80/window harness floor. So this is
    reported as descriptive context next to the daily variant; the harness
    verdict is rendered on the daily variant.
    """
    records: list[dict] = []
    gross_rets: list[float] = []
    net_rets: list[float] = []
    cost_frac = ROUND_TRIP_COST_BPS / 10000.0

    for ticker, series in data.items():
        price = series.get("price", {})
        dates = sorted(price)
        # Pick the last trading day per (year, month) as the rebalance date.
        ym_to_last: dict[str, str] = {}
        for d in dates:
            ym = d[:7]
            ym_to_last[ym] = d
        month_dates = sorted(ym_to_last.values())
        for k in range(LOOKBACK_MONTHS, len(month_dates) - 1):
            d_entry = month_dates[k]
            d_lookback = month_dates[k - LOOKBACK_MONTHS]
            d_exit = month_dates[k + 1]
            p_lookback = price.get(d_lookback)
            p_entry = price.get(d_entry)
            p_exit = price.get(d_exit)
            if not (p_lookback and p_entry and p_exit) or p_lookback <= 0:
                continue
            tr = p_entry / p_lookback - 1.0
            direction = 1 if tr > 0 else (-1 if tr < 0 else 0)
            if direction == 0:
                continue
            raw_ret = p_exit / p_entry - 1.0
            signed_ret = raw_ret * direction
            net_ret = signed_ret - cost_frac
            status = "WON" if signed_ret > 0 else "LOST"
            gross_rets.append(signed_ret)
            net_rets.append(net_ret)
            records.append({
                "status": status,
                "resolved_at": d_exit,
                "entry_date": d_entry,
                "signal_date": d_entry,
                "timestamp": d_entry,
                "signed_ret": round(signed_ret, 8),
                "net_ret": round(net_ret, 8),
                "direction": direction,
                "instrument": ticker,
                "asset_class": ASSET_CLASS_MAP.get(ticker, "FUTURES"),
            })

    return {
        "records": records,
        "gross_rets": gross_rets,
        "net_rets": net_rets,
    }


# ===========================================================================
# Post-cost gate
# ===========================================================================
def cost_survival(gross_rets: list[float], net_rets: list[float]) -> dict:
    """Post-cost gate: does net per-trade edge keep >= 60% of gross?"""
    if not gross_rets:
        return {"passes": False, "reason": "no trades"}
    gross_bps = statistics.fmean(gross_rets) * 10000.0
    net_bps = statistics.fmean(net_rets) * 10000.0
    survival = 0.0 if gross_bps <= 0 else net_bps / gross_bps
    return {
        "gross_edge_bps": round(gross_bps, 4),
        "net_edge_bps": round(net_bps, 4),
        "cost_bps": ROUND_TRIP_COST_BPS,
        "cost_survival_pct": round(survival * 100, 2),
        "passes": bool(gross_bps > 0 and survival >= COST_SURVIVAL_MIN),
    }


def pooled_wr(records: list[dict], net: bool = False):
    """Pooled win-rate (gross or net of cost)."""
    key = "net_ret" if net else "signed_ret"
    rs = [r.get(key) for r in records if r.get(key) is not None]
    if not rs:
        return None
    return round(sum(1 for r in rs if r > 0) / len(rs), 4)


# ===========================================================================
# Harness wiring — edge_stability_harness imported UNMODIFIED
# ===========================================================================
def harness_verdict(records: list[dict], window_days: int = WINDOW_DAYS) -> dict:
    """Run synthetic records through edge_stability_harness, patching only its
    loader for this call. The harness logic — windowing, eff, is_admissible()
    threshold — is reused VERBATIM and UNMODIFIED.
    """
    orig_load = harness._load
    try:
        harness._load = lambda: records  # type: ignore[assignment]
        verdict = harness.evaluate(HARNESS_FIELD, window_days)
        admissible = harness.is_admissible(HARNESS_FIELD, window_days)
    finally:
        harness._load = orig_load  # type: ignore[assignment]
    verdict["admissible_via_is_admissible"] = admissible
    return verdict


# ===========================================================================
# Report
# ===========================================================================
def render_report(res: dict) -> str:
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds")
    recs = res.get("records", [])
    h = res.get("harness", {})
    cg = res.get("cost_gate", {})
    offline = res.get("any_offline", False)
    aqr = res.get("aqr_sanity", {})
    adm_harness = h.get("admissible", False)
    cost_pass = cg.get("passes", False)
    scored = h.get("windows_scored", 0)

    if offline or scored < harness.MIN_STABLE_WINDOWS:
        verdict = "UNTESTED-data-gap"
    elif adm_harness and cost_pass:
        verdict = "ADMISSIBLE"
    else:
        verdict = "REJECTED"

    out = [
        "# H-035 FUTURES Time-Series Momentum Diversified — 2026-05-19",
        "",
        f"_Generated {ts} by `tools/h035_futures_tsmom_research.py`._",
        "",
        f"## VERDICT: **{verdict}**",
        "",
        "**Status: OPT-IN RESEARCH SIDECAR. HARNESS-GATED. FREE-DATA. NO "
        "PRODUCTION WIRING.** No caller in `quality_gates.py`, "
        "`dashboard_generator.py`, `passes_active_gate`, `calculate_smart_score`, "
        "or any pick-generation / scoring path. Fetches free public futures "
        "data, runs the backtest, gates through `edge_stability_harness` "
        "(imported UNMODIFIED), writes this report — nothing else. "
        "Multi-AI-D-path consensus (Grok + Claude swarm 2026-05-19).",
        "",
        "## The signal (Moskowitz-Ooi-Pedersen 2012)",
        "",
        "Time-series momentum across a diversified liquid-futures universe — "
        "equity indices, currencies, commodities, government bonds. The sign "
        "of each contract's trailing 12-month total return is the direction; "
        "each leg is vol-scaled to a target ~10% annualised volatility; the "
        "diversified book is equal-weighted; rebalance monthly.",
        "",
        f"- look-back: **{LOOKBACK_MONTHS} months ({LOOKBACK_DAYS} trading days)**",
        f"- vol target per leg: **{TARGET_VOL_ANNUAL*100:.1f}% annualised** "
        f"(realised over the trailing {VOL_LOOKBACK_DAYS}d, strictly past)",
        f"- conviction magnitude (harness score field): |z-score of trailing "
        f"return against its own {Z_ROLL}-obs strictly-past distribution|",
        f"- universe: {', '.join(UNIVERSE_FULL)}",
        "",
        "**Causal mechanism:** Moskowitz, Ooi & Pedersen 2012 documented the "
        "TSMOM premium across 58 contracts spanning equities, FX, commodities, "
        "bonds (1985-2009). The premium is consistent with under-reaction to "
        "information and delayed over-reaction by uninformed momentum traders. "
        "The diversification across uncorrelated futures classes is the moat — "
        "a single-class TSMOM book is noisier. A sound prior is NOT an edge; "
        "only the harness + cost verdict counts.",
        "",
        "## Construction (legitimate density pattern from H-008/H-019/H-026/H-027/H-028)",
        "",
        "- **FULL continuous-position diversified book** — one resolved record "
        "per (contract, day) on the daily-rebalance variant. NO sparse event "
        "filter, NO threshold. The daily variant is what feeds the harness; "
        "a parallel classical monthly-rebalance MOP book is computed for "
        "descriptive context only (monthly records are too sparse for the "
        "14-day harness window).",
        "- **Vol-scaling per leg** — weight = `TARGET_VOL / realised_vol_60d`, "
        "capped at 5x to keep extreme low-vol periods from blowing up. The "
        "record stores `weight` and `signed_ret = raw_ret * direction * weight`.",
        "- **Strict no-look-ahead** — trailing return and realised vol are "
        "computed from prices STRICTLY at or before day D; the position is "
        "entered at close(D) and exited at close(D+1).",
        "",
        "## Data (free sources only — yfinance =F continuous-front)",
        "",
        "- **PRIMARY (verdict input):** yfinance `=F` continuous-front-month "
        "tickers (free, no key). The Claude swarm verdict (2026-05-19) was "
        "explicit: *\"AQR publishes factor returns but NOT constituent "
        "positions. You need to reimplement using yfinance =F tickers.\"* "
        "So the yfinance reconstruction IS the real implementation.",
        "- **OPTIONAL SANITY CHECK:** AQR's published monthly TSMOM factor "
        "return series, manually downloaded from "
        "<https://www.aqr.com/Insights/Datasets/Time-Series-Momentum-Factors-Monthly> "
        "into `tools/cache/aqr_tsmom_monthly.csv` (or `.xlsx`). When present, "
        "it is reported alongside the reconstruction as a smell-test — it "
        "CANNOT drive the harness (no per-contract positions).",
        f"- **Sample:** {len(recs)} contract-day resolved records.",
        (
            "- **OFFLINE NOTE:** one or more contracts fell back to a "
            "labelled synthetic series because yfinance was unreachable. The "
            "verdict is therefore UNTESTED-data-gap — this run did NOT measure "
            "a real edge."
            if offline else
            "- All contracts sourced from live free yfinance data."
        ),
        "",
    ]

    if aqr:
        out += ["## AQR published TSMOM sanity check (NOT the verdict)",
                "",
                f"- AQR monthly returns loaded: **{aqr.get('n_months', 0)} months**",
                f"- AQR mean monthly return: {aqr.get('mean_monthly_pct', 0):.3f}%",
                f"- AQR ann. return (12 * mean): {aqr.get('annualised_pct', 0):.2f}%",
                "- Note: AQR's series is the factor-return aggregate. It does "
                "NOT carry per-contract positions, so it cannot drive the "
                "harness. The yfinance reconstruction is the verdict input. "
                "If the published series goes one way and our reconstruction "
                "goes another, the *reconstruction* is the smell — not the "
                "academic series.",
                ""]

    pi = res.get("per_instrument", {})
    if pi:
        out += ["| Contract | class | records | wins | gross WR |",
                "|---|---|---|---|---|"]
        for k, v in pi.items():
            wr = f"{v['wr']*100:.1f}%" if v.get("wr") is not None else "n/a"
            out.append(f"| {k} | {v.get('asset_class','FUTURES')} | "
                       f"{v.get('n',0)} | {v.get('wins',0)} | {wr} |")
        out.append("")

    out += ["## Harness verdict (THE gate — `edge_stability_harness`, UNMODIFIED)",
            ""]
    if "per_window_eff" in h:
        effs = " ".join(
            (f"{e['eff']:+.2f}" if e.get("eff") is not None else "n/a")
            for e in h["per_window_eff"])
        out += [
            f"- per-window eff (new->old): `{effs}`",
            f"- windows scored: {h.get('windows_scored')}  "
            f"(strong {h.get('windows_strong')}: "
            f"+{h.get('strong_positive')}/-{h.get('strong_negative')})",
            f"- `is_admissible()`: {h.get('admissible_via_is_admissible')}",
            f"- harness reason: {h.get('reason', 'n/a')}",
            "",
        ]
    else:
        out += [f"- {h.get('reason', 'n/a')}", ""]

    out += ["## Post-cost gate (10bps round-trip; net must keep >= 60% of gross)",
            ""]
    if cg:
        out += [
            f"- gross edge: **{cg.get('gross_edge_bps')} bps/trade**",
            f"- net edge (after {cg.get('cost_bps')}bps round-trip): "
            f"**{cg.get('net_edge_bps')} bps/trade**",
            f"- cost-survival: **{cg.get('cost_survival_pct')}%** "
            f"(floor {int(COST_SURVIVAL_MIN*100)}%)",
            f"- post-cost gate: **{'PASS' if cost_pass else 'FAIL'}**",
            "",
        ]
    out += [
        f"- pooled gross WR: {(res.get('pooled_wr_gross') or 0)*100:.1f}%",
        f"- pooled net WR: {(res.get('pooled_wr_net') or 0)*100:.1f}%",
        "",
        "## Honest conclusion",
        "",
    ]
    if verdict == "ADMISSIBLE":
        out += [
            "**H-035 cleared `edge_stability_harness` AND the post-cost gate.** "
            "Against a poor base rate (8-11 prior kills) this is a **research "
            "candidate, not a green light.** Before any wiring it still needs: "
            "fresh out-of-sample re-test on data collected after the "
            "`data_sample_lock`, a deflated-Sharpe / SPA multiple-testing "
            "correction (TSMOM has been mined for decades — multiple-testing "
            "haircut is non-trivial), an examination of the post-2010 sub-period "
            "performance separately (the published TSMOM premium has been "
            "criticised as decaying since the GFC), and operator review. No "
            "wiring here.",
        ]
    elif verdict == "REJECTED":
        out += [
            "**H-035 was TESTED and REJECTED.** The diversified continuous-"
            "position construction gave the harness real window density, so it "
            "rendered a genuine verdict. " + (
                "The harness eff sign splits across windows — in-sample "
                "separation that flips sign across regimes (the same failure "
                "mode that killed H-014, H-038, method_a_score). "
                if not adm_harness else
                "The harness sign is stable but the post-cost gate fails. "
            ) + "Do NOT re-test this construction on this sample. Nothing is "
            "wired or sized.",
        ]
    else:
        out += [
            "**H-035 is UNTESTED — data-gap, explicitly NOT a pass.** " + (
                "yfinance was unreachable for one or more contracts, so the "
                "run used a labelled synthetic series. "
                if offline else
                f"The harness scored only {scored} fourteen-day window(s); it "
                f"needs >= {harness.MIN_STABLE_WINDOWS}. "
            ) + "This is an honest non-verdict — not a pass and not a clean "
            "fail. Re-run only with the dense free yfinance =F tape over the "
            "full ~25-year history.",
        ]
    out += [
        "",
        "## next_step",
        "",
        (
            "H-035 cleared both gates — escalate for out-of-sample re-test on "
            "data collected after the data_sample_lock, deflated-Sharpe / SPA "
            "correction (TSMOM has been mined for decades), and operator review "
            "before any wiring."
            if verdict == "ADMISSIBLE" else
            "Harness kill — diversified TSMOM does not yield an admissible "
            "futures edge on this sample. Do NOT re-test this construction on "
            "this data."
            if verdict == "REJECTED" else
            "Re-run with the dense yfinance =F tape (network was unavailable "
            "or harness density was below the 14-day-window floor). The "
            "monthly-rebalance MOP variant is too sparse for the harness; the "
            "daily-rebalance variant is the harness-input book."
        ),
        "",
    ]
    return "\n".join(out)


# ===========================================================================
# Orchestration
# ===========================================================================
def run(quick: bool, refresh: bool) -> dict:
    tickers = UNIVERSE_QUICK if quick else UNIVERSE_FULL
    data = load_market_data(tickers, refresh)

    bt_daily = backtest_continuous(data)
    bt_monthly = backtest_monthly(data)
    recs = bt_daily.get("records", [])
    cg = cost_survival(bt_daily.get("gross_rets", []), bt_daily.get("net_rets", []))

    # AQR sanity check (optional).
    aqr_series = load_aqr_tsmom_optional()
    aqr_sanity: dict = {}
    if aqr_series:
        vals = list(aqr_series.values())
        if vals:
            mean_m = statistics.fmean(vals)
            aqr_sanity = {
                "n_months": len(vals),
                "mean_monthly_pct": round(mean_m * 100, 4),
                "annualised_pct": round(mean_m * 12 * 100, 2),
            }

    res = {
        "records": recs,
        "per_instrument": bt_daily.get("per_instrument", {}),
        "cost_gate": cg,
        "any_offline": bt_daily.get("any_offline", False),
        "pooled_wr_gross": pooled_wr(recs, net=False),
        "pooled_wr_net": pooled_wr(recs, net=True),
        "monthly_n_records": len(bt_monthly.get("records", [])),
        "aqr_sanity": aqr_sanity,
    }
    if len(recs) < harness.MIN_WINDOW_N:
        res["harness"] = {
            "admissible": False,
            "reason": f"INSUFFICIENT DATA — {len(recs)} records, harness needs "
                      f">= {harness.MIN_WINDOW_N}/window",
            "windows_scored": 0,
        }
    else:
        res["harness"] = harness_verdict(recs)
    return res


def json_summary(res: dict) -> dict:
    h = res.get("harness", {})
    cg = res.get("cost_gate", {})
    scored = h.get("windows_scored", 0)
    offline = res.get("any_offline", False)
    if offline or scored < harness.MIN_STABLE_WINDOWS:
        verdict = "UNTESTED-data-gap"
    elif h.get("admissible") and cg.get("passes"):
        verdict = "ADMISSIBLE"
    else:
        verdict = "REJECTED"
    return {
        "hypothesis": "H-035",
        "family": "futures_time_series_momentum_diversified",
        "asset_class": "FUTURES",
        "verdict": verdict,
        "offline_synthetic": offline,
        "n_records": len(res.get("records", [])),
        "monthly_n_records": res.get("monthly_n_records"),
        "windows_scored": scored,
        "windows_strong": h.get("windows_strong"),
        "strong_positive": h.get("strong_positive"),
        "strong_negative": h.get("strong_negative"),
        "per_window_eff": [e.get("eff") for e in h.get("per_window_eff", [])],
        "harness_sign": h.get("sign"),
        "is_admissible": h.get("admissible_via_is_admissible", h.get("admissible")),
        "pooled_wr_gross": res.get("pooled_wr_gross"),
        "pooled_wr_net": res.get("pooled_wr_net"),
        "gross_edge_bps": cg.get("gross_edge_bps"),
        "net_edge_bps": cg.get("net_edge_bps"),
        "cost_survival_pct": cg.get("cost_survival_pct"),
        "cost_gate_passes": cg.get("passes"),
        "harness_reason": h.get("reason"),
        "aqr_sanity": res.get("aqr_sanity", {}),
    }


def main() -> int:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--quick", action="store_true",
                    help="5-contract smoke set for a fast run")
    ap.add_argument("--json", dest="as_json", action="store_true")
    ap.add_argument("--refresh-cache", action="store_true",
                    help="re-fetch futures data, ignoring the cache")
    ap.add_argument("--out", type=Path,
                    default=ROOT / "reports" /
                    "h035_futures_tsmom_research_2026-05-19.md")
    args = ap.parse_args()

    res = run(args.quick, args.refresh_cache)
    summary = json_summary(res)

    if args.as_json:
        print(json.dumps(summary, indent=2, default=str))
        return 0

    report = render_report(res)
    args.out.write_text(report, encoding="utf-8")
    print(f"# wrote {args.out}", file=sys.stderr)
    print(report)
    print("\n# JSON SUMMARY")
    print(json.dumps(summary, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
